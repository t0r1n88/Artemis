"""
Скрипт для обработки создания отчетов по площадям леса
"""
import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import time
import datetime
pd.options.mode.chained_assignment = None  # default='warn'
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller
    Функция чтобы логотип отображался"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def select_end_folder():
    """
    Функция для выбора конечной папки куда будут складываться итоговые файлы
    :return:
    """
    global path_to_end_folder
    path_to_end_folder = filedialog.askdirectory()


def select_file_data_xlsx():
    """
    Функция для выбора файла с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global file_data_xlsx
    # Получаем путь к файлу
    file_data_xlsx = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

def select_file_reestr_purpose():
    global reest_upp_purpose
    # Получаем путь к файлу
    reest_upp_purpose = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

"""
Функции для работы проверки наличия записи участка УППП
"""
def select_file_params_presense():
    """
    Функция для выбора файла с номерами колонок по которым будет вестись сравнение
    :return: Путь к файлу с параметрами
    """
    global file_params_presense
    # Получаем путь к файлу
    file_params_presense = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

def select_file_reestr_presense():
    """
    Функция для выбора файла реестра. Отдельные функции для таких простых вещей сделаны специально чтобы другому человеку
    было легче разобраться
    """
    global file_reestr_presense
    file_reestr_presense = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

def select_file_statement_presense():
    """
    Функция для выбора файла ведомости записи  в которой нужно проверить в реестре на наличие
    """
    global file_statement_presense
    file_statement_presense = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

"""
Подготовительные функции для переноса данных из таблицы 3 в реестр УПП
"""
def select_header_reestr():
    """
    Функция для выбора файла ведомости записи  в которой нужно проверить в реестре на наличие
    """
    global header_reestr
    header_reestr = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

def select_reestr_transfer3():
    """
    Функция для выбора файла ведомости записи  в которой нужно проверить в реестре на наличие
    """
    global file_transfer_reestr
    file_transfer_reestr = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_table3_transfer3():
    """
    Функция для выбора файла ведомости записи  в которой нужно проверить в реестре на наличие
    """
    global file_transfer_to_upp
    file_transfer_to_upp = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))



def processing_presense_reestr():
    """
    Функция проверяеющая наличие записей из ведомости в в реестре УПП
    """
    try:
        # Считываем из файлов только те колонки по которым будет вестись сравнение
        first_df = pd.read_excel(file_reestr_presense,
                                 skiprows=8, usecols=[0, 1, 2, 3, 4], keep_default_na=False)
        second_df = pd.read_excel(file_statement_presense, usecols=[0, 1, 2, 3, 4], keep_default_na=False)

        # Приводим к строковому формату названия колонок первого датафрейма
        first_df.columns = list(map(str, list(first_df.columns)))
        # на всякий случай очищаем от пробельных символов
        first_df.columns = list(map(lambda x: x.replace(" ", ""), list(first_df.columns)))

        # заменяем на цифры названия колонок во втором датафрейме
        second_df.columns = ['1', '2', '3', '4', '5']

        # Приводим датафреймы к строковому виду
        first_df = first_df.astype(str)
        second_df = second_df.astype(str)

        # Очищаем от
        first_df.replace(r'^\s*$', 'Отсутствует', regex=True, inplace=True)
        second_df.replace(r'^\s*$', 'Отсутствует', regex=True, inplace=True)

        # так как мы заранее знаем сколько и какие колонки у нас есть то просто создаем список
        params_columns = [0, 1, 2, 3, 4]
        # Конвертируем нужные нам колонки в str
        convert_columns_to_str(first_df, params_columns)
        convert_columns_to_str(second_df, params_columns)

        """
        Соответствие названий колонок используемым в программе номерам колонок
        Лесничество -1
        Участковое лесничество- 2
        Урочище - 3
        Номер лесного квартала -4
        Номер лесотаксационного выдела -5

        """

        # Готовим 4 и 5 колонки чтобы они были разделеныпри склеивании
        first_df['4'] = first_df['4'].apply(lambda x: 'кв.' + x)  # Добавляем разделитель квартал
        first_df['5'] = first_df['5'].apply(lambda x: 'в.' + x)  # Добавляем разделитель выдел

        # То же самое для второго датафрейма
        second_df['4'] = second_df['4'].apply(lambda x: 'кв.' + x)  # Добавляем разделитель квартал
        second_df['5'] = second_df['5'].apply(lambda x: 'в.' + x)  # Добавляем разделитель выдел

        # Создаем в каждом датафрейме колонку с айди путем склеивания всех нужных колонок в одну строку
        first_df['ID'] = first_df.iloc[:, params_columns].sum(axis=1)
        second_df['ID'] = second_df.iloc[:, params_columns].sum(axis=1)

        # Обрабатываем дубликаты

        first_df.drop_duplicates(subset=['ID'], keep='last', inplace=True)  # Удаляем дубликаты из датафрейма

        second_df.drop_duplicates(subset=['ID'], keep='last', inplace=True)  # Удаляем дубликаты из датафрейма

        # Создаем документ
        wb = openpyxl.Workbook()
        # создаем листы
        ren_sheet = wb['Sheet']
        ren_sheet.title = 'Итог'

        # Создаем датафрейм
        itog_df = pd.merge(first_df, second_df, how='outer', left_on=['ID'], right_on=['ID'],
                           indicator=True)

        # Отфильтровываем значения both,right
        out_df = itog_df[(itog_df['_merge'] == 'both') | (itog_df['_merge'] == 'right_only')]

        out_df.rename(
            columns={'_merge': 'Присутствие в реестре УПП', '1_x': 'Реестр Лесничество', '2_x': 'Реестр Уч.лесничество',
                     '3_x': 'Реестр Урочище',
                     '4_x': 'Реестр Квартал', '5_x': 'Реестр Выдел', '1_y': 'Ведомость Лесничество',
                     '2_y': 'Ведомость Уч.Лесничество',
                     '3_y': 'Ведомость Урочище', '4_y': 'Ведомость Квартал', '5_y': 'Ведомость Выдел'}, inplace=True)

        out_df['Присутствие в реестре УПП'] = out_df['Присутствие в реестре УПП'].apply(
            lambda x: 'Имеется в реестре' if x == 'both' else 'Отсутствует в реестре')

        # Получаем текущую дату
        current_time = time.strftime('%H_%M_%S %d.%m.%Y')
        # Сохраняем отчет
        # Для того чтобы увеличить ширину колонок для удобства чтения используем openpyxl
        wb = openpyxl.Workbook()  # Создаем объект
        # Записываем результаты
        for row in dataframe_to_rows(out_df, index=False, header=True):
            wb['Sheet'].append(row)

        # Форматирование итоговой таблицы
        # Ширина колонок
        wb['Sheet'].column_dimensions['A'].width = 15
        wb['Sheet'].column_dimensions['B'].width = 20
        wb['Sheet'].column_dimensions['C'].width = 10
        wb['Sheet'].column_dimensions['G'].width = 20
        wb['Sheet'].column_dimensions['H'].width = 20
        wb['Sheet'].column_dimensions['F'].width = 50
        wb['Sheet'].column_dimensions['L'].width = 30
        # Перенос строк для заголовков
        wb['Sheet']['D1'].alignment = Alignment(wrap_text=True)
        wb['Sheet']['E1'].alignment = Alignment(wrap_text=True)
        wb['Sheet']['F1'].alignment = Alignment(wrap_text=True)
        wb['Sheet']['G1'].alignment = Alignment(wrap_text=True)
        wb['Sheet']['H1'].alignment = Alignment(wrap_text=True)
        wb['Sheet']['I1'].alignment = Alignment(wrap_text=True)
        wb['Sheet']['J1'].alignment = Alignment(wrap_text=True)
        wb['Sheet']['K1'].alignment = Alignment(wrap_text=True)
        wb['Sheet']['H1'].alignment = Alignment(wrap_text=True)

        wb.save(
            f'{path_to_end_folder}/Сравнение УПП с другими ведомостями на наличие участков или их отсутствие от  {current_time}.xlsx')

    except ValueError as e:
        messagebox.showerror('Артемида 1.8',
                             f'Не найдена колонка или лист {e.args}\nДанные в файле должны находиться на листе с названием Реестр УПП'
                             f'\nКолонки 1-8 должны иметь названия: Лесничество,Участковое лесничество,Урочище ,Номер лесного квартала,\n'
                             f'Номер лесотаксационного выдела,Площадь лесотаксационного выдела, га,Обозначение части лесотаксационного выдела (лесопатологического выдела), га ,'
                             f'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га')
    except NameError:
        messagebox.showerror('Артемида 1.8', f'Выберите файл с данными и конечную папку')
    except PermissionError:
        messagebox.showerror('Артемида 1.8', f'Закройте файлы с созданными раньше отчетами!!!')
    except FileNotFoundError:
        messagebox.showerror('Артемида 1.8', f'Проверьте наличие указанных файлов')
    except MemoryError:
        messagebox.showerror('Артемида 1.8', f'Слишком большая таблица!!!\nПроверьте размер таблицы!!!\nНажмите CTRL+END для проверки'
                                             f'Пересоздайте файл без пустых строк и столбцов')
    else:
        messagebox.showinfo('Артемида 1.8', 'Работа программы успешно завершена!!!')

def convert_to_int_transfer(cell):
    """
    Функция для конвертации в int
    """
    try:
        return int(cell)
    except ValueError:
        return 0

def combine(x):
    # Функция для группировки всех значений в строку разделенную ;
    return ';'.join(x)


def check_unique(x):
    # Функция для нахождения разночтений в значениях
    temp_lst = x.split(';')
    temp_set = set(temp_lst)
    if 'nan' in temp_set:
        return 'Не заполнены значения площади лесотаксационного выдела!!!'
    else:
        return 'Все в порядке' if len(temp_set) == 1 else 'Ошибка!!!'


def main_check_unique(x):
    # Функция для проверки корректности заполнения площади выдела
    temp_str = ';'.join(x)  # Склеиваем все значения
    temp_lst = temp_str.split(';')  # Создаем список разбивая по ;
    temp_set = set(temp_lst)  # Превращаем в множество
    if len(temp_set) > 1:  # Если длина множества больше 1 то есть погрешности
        return 0
    elif 'nan' in temp_set:  # если есть нан то не заполнены площади выдела
        return 0
    else:  # Если все в порядке то возвращаем единственный элемент списка
        return convert_to_float(temp_lst[0])


def convert_to_float(x):
    """
    Функция для конвертирования строки в флоат при ошибке возвращает 0
    :param x: строка
    :return:
    """
    try:
        return float(x)
    except ValueError:
        return 0

def clean_purpose_column(x):
    """
    Функция для извлечения значений из столбца целевого назначения для того чтобы можно было
    найти все значения равные 1 и сопоставить со значением в категории
    """
    temp_lst = x.split(';') # Создаем список разделя строку по ;
    temp_set = set(temp_lst) # Превращаем во множество

    if len(temp_set) == 1:
        temp_value = list(temp_set)[0] # получаем единственное значение
        if temp_value == 'nan':
            return 0
        try:
            value_purpose = float(temp_value) # конвертируем в число
            return value_purpose
        except ValueError:
            return 0
    else:
        return 0

def prepare_column_purpose_category(df,name_columns):
    """
    Функция для предобработки колонок с целевым назначением и категорией лесов
    Нужно очистить от пробелов, nan, сконвертировать во флоат,инт и снова в строку
    :param df: датафрейм содержащий в себе реестр
    :param name_columns: название обрабаытваемой колонки
    """
    try:
        # Приводим колонку к типу str чтобы очистить от лишних символов и заменить пустые вещи на нули
        df[name_columns] = df[name_columns].astype(str)
        df[name_columns] = df[name_columns].apply(lambda x: x.replace('nan', '0'))
        df[name_columns] = df[name_columns].apply(lambda x: x.replace(' ', '0'))
        df[name_columns] = df[name_columns].apply(lambda x: x.strip())
        # конвертируем во флоат, затем в инт чтобы потом в строке не было значений с дробной частью
        df[name_columns] = df[name_columns].apply(convert_to_float)
        df[name_columns] = df[name_columns].astype(int)
        df[name_columns] = df[name_columns].astype(str)
    except KeyError as e:
            messagebox.showerror('Артемида 1.8',f'Не найдена колонка {e.args} Проверьте файл на наличие этой колонки')
    except ValueError as e:
        messagebox.showerror('Артемида 1.8', f'Возникла ошибка при обработке значения {e.args}\n'
                                             f'в колонках целевого назначения и категории должны быть только цифры!')

def convert_columns_to_str(df, number_columns):
    """
    Функция для конвертации указанных столбцов в строковый тип и очистки от пробельных символов в начале и конце
    """

    for column in number_columns:  # Перебираем список нужных колонок
        df.iloc[:, column] = df.iloc[:, column].astype(str)
        # Очищаем колонку от пробельных символов с начала и конца
        df.iloc[:, column] = df.iloc[:, column].apply(lambda x: x.strip())
        df.iloc[:, column] = df.iloc[:, column].apply(lambda x: x.replace(' ', ''))


def convert_params_columns_to_int(lst):
    """
    Функция для конвератации значений колонок которые нужно обработать.
    Очищает от пустых строк, чтобы в итоге остался список из чисел в формате int
    """
    out_lst = []  # Создаем список в который будем добавлять только числа
    for value in lst:  # Перебираем список
        try:
            # Обрабатываем случай с нулем, для того чтобы после приведения к питоновскому отсчету от нуля не получилась колонка с номером -1
            number = int(value)
            if number != 0:
                out_lst.append(value)  # Если конвертирования прошло без ошибок то добавляем
            else:
                continue
        except:  # Иначе пропускаем
            continue
    return out_lst

def prepare_column_to_int(cell):
    """
    Функция для обработки ячеек в инт, чтобы они в экселе потом не отображались как строки
    """
    try:
        float_value = float(cell)
        int_value = int(float_value)
        return int_value
    except ValueError:
        return cell

def clean_column(df,lst_column):
    """
    Функция для очистки от колонки от пустых значений
    """
    for column in lst_column:
        df[column].fillna('Отсутствует',inplace=True)
        df[column] = df[column].astype(str) # Приводим к строковому виду
        df[column] = df[column].apply(lambda x:x.replace('nan','Отсутствует'))
        df[column] = df[column].apply(lambda x:x.strip()) # очищаем от пробельных символов(на случай если в ячейке стоит просто пробел или несколько пробелов)
        df[column] = df[column].apply(lambda x:x if x else 'Отсутствует') # Если пустая строка то заменяем на значение Не заполнено

def processing_report_square_wood():
    """
    Фугкция для обработки данных
    :return:
    """
    try:
        # Получаем значение чекбокса региона
        region = group_rb_region_report_square.get()
        use_cols = list(range(25))

        df = pd.read_excel(file_data_xlsx, skiprows=8, usecols=use_cols)
        # Приводим названия колонок к строковому виду, чтобы избежать возможных проблем с названиями колонок
        df.columns = list(map(str, list(df.columns)))
        # на всякий случай очищаем от пробельных символов
        df.columns = list(map(lambda x: x.replace(" ", ""), list(df.columns)))

        """
        Соответствие названий колонок используемым в программе номерам колонок
        Лесничество -1
        Участковое лесничество- 2
        Урочище - 3
        Номер лесного квартала -4
        Номер лесотаксационного выдела -5
        Площадь лесотаксационного выдела, га - 6
        Площадь лесотаксационного выдела или его части (лесопатологического выдела), га - 8
        Год лесоустройства -11



        """

        df.rename(
            columns={'1': 'Лесничество', '2': 'Участковое лесничество', '3': 'Урочище', '4': 'Номер лесного квартала',
                     '5': 'Номер лесотаксационного выдела',
                     '6': 'Площадь лесотаксационного выдела, га',
                     '8': 'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га',
                     '11': 'Год лесоустройства'}, inplace=True)

        # заполняем пропущенные места
        clean_column(df, ['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                          'Номер лесотаксационного выдела', 'Год лесоустройства'])

        # Бурятия
        if region == 0:
            # СОздаем проверочный файл для проверки правильности ввода плошади выдела
            check_df = df.copy(deep=True)

            # Меняем тип столбца на строку чтобы создать строку включающую в себя все значения разделенные ;
            check_df['Площадь лесотаксационного выдела, га'] = check_df['Площадь лесотаксационного выдела, га'].astype(
                str)
            #

            checked_pl = check_df.groupby(['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                           'Номер лесотаксационного выдела']).agg(
                {'Площадь лесотаксационного выдела, га': combine})

            # Применяем функцию првоеряющую количество уникальных значений в столбце, если больше одного то значит есть ошибка в данных
            checked_pl['Контроль совпадения площади выдела'] = checked_pl['Площадь лесотаксационного выдела, га'].apply(
                check_unique)

            # переименовывам колонку
            checked_pl.rename(
                columns={'Площадь лесотаксационного выдела, га': 'Все значения площади для указанного выдела'},
                inplace=True)
            # Извлекаем индексы в колонки
            checked_pl = checked_pl.reset_index()

            # конвертируем в инт чтобы корректно отображалось
            checked_pl['Номер лесного квартала'] = checked_pl['Номер лесного квартала'].apply(prepare_column_to_int)
            checked_pl['Номер лесотаксационного выдела'] = checked_pl['Номер лесотаксационного выдела'].apply(
                prepare_column_to_int)

            # Сортируем
            checked_pl.sort_values(by=['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                       'Номер лесотаксационного выдела'], inplace=True)

            # Получаем текущую дату
            current_time = time.strftime('%H_%M_%S %d.%m.%Y')
            # Сохраняем отчет
            # Для того чтобы увеличить ширину колонок для удобства чтения используем openpyxl
            wb = openpyxl.Workbook()  # Создаем объект
            # Записываем результаты
            for row in dataframe_to_rows(checked_pl, index=False, header=True):
                wb['Sheet'].append(row)

            # Форматирование итоговой таблицы
            # Ширина колонок
            wb['Sheet'].column_dimensions['A'].width = 15
            wb['Sheet'].column_dimensions['B'].width = 20
            wb['Sheet'].column_dimensions['C'].width = 36
            wb['Sheet'].column_dimensions['F'].width = 15
            wb['Sheet'].column_dimensions['G'].width = 15
            # Перенос строк для заголовков
            wb['Sheet']['D1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['E1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['F1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['G1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['H1'].alignment = Alignment(wrap_text=True)

            wb.save(
                f'{path_to_end_folder}/Бурятия Проверка правильности ввода площадей лесотаксационного выдела {current_time}.xlsx')

            # Основной отчет
            # Готовим колонки к группировке
            df['Площадь лесотаксационного выдела, га'] = df['Площадь лесотаксационного выдела, га'].astype(str)

            df['Площадь лесотаксационного выдела, га'] = df['Площадь лесотаксационного выдела, га'].apply(
                lambda x: x.replace(',', '.'))

            df['Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'] = df[
                'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'].astype(str)

            df['Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'] = df[
                'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'].apply(
                lambda x: x.replace(',', '.'))

            df['Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'] = df[
                'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'].apply(
                convert_to_float)

            # Группируем
            group_df = df.groupby(['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                   'Номер лесотаксационного выдела']).agg(
                {'Площадь лесотаксационного выдела, га': main_check_unique,
                 'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га': 'sum'})

            group_df = group_df.astype(str)

            # переименовываем колонку
            group_df.rename(columns={
                'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га': 'Используемая площадь лесотаксационного выдела, га'},
                inplace=True)

            # Извлекаем индексы в колонки
            group_df = group_df.reset_index()

            group_df['Площадь лесотаксационного выдела, га'] = group_df['Площадь лесотаксационного выдела, га'].apply(
                convert_to_float)
            group_df['Используемая площадь лесотаксационного выдела, га'] = group_df[
                'Используемая площадь лесотаксационного выдела, га'].apply(convert_to_float)

            # Округляем до 3 знаков для корректного сравнения
            group_df['Площадь лесотаксационного выдела, га'] = np.round(
                group_df['Площадь лесотаксационного выдела, га'], decimals=3)
            group_df['Используемая площадь лесотаксационного выдела, га'] = np.round(
                group_df['Используемая площадь лесотаксационного выдела, га'], decimals=3)

            # Создаем колонку для контроля
            group_df['Контроль площади используемого надела'] = group_df['Площадь лесотаксационного выдела, га'] < \
                                                                group_df[
                                                                    'Используемая площадь лесотаксационного выдела, га']

            group_df['Контроль площади используемого надела'] = group_df['Контроль площади используемого надела'].apply(
                lambda x: 'Превышение используемой площади выдела!!!' if x is True else 'Все в порядке')

            # Изменяем состояние колонки если площадь всего выдела равна 0
            group_df['Контроль правильности ввода площади лесотаксационного выдела'] = group_df[
                'Площадь лесотаксационного выдела, га'].apply(
                lambda
                    x: 'Площадь лесотаксационного выдела равна нулю или  обнаружены разные значения площади выдела !!!' if x == 0 else 'Площади лесотаксационного выдела не отличаются друг от друга')

            # конвертируем в инт чтобы корректно отображалось
            group_df['Номер лесного квартала'] = group_df['Номер лесного квартала'].apply(prepare_column_to_int)
            group_df['Номер лесотаксационного выдела'] = group_df['Номер лесотаксационного выдела'].apply(
                prepare_column_to_int)

            # Сортируем
            group_df.sort_values(by=['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                     'Номер лесотаксационного выдела'], inplace=True)

            # Сохраняем отчет
            # Для того чтобы увеличить ширину колонок для удобства чтения используем openpyxl
            wb = openpyxl.Workbook()  # Создаем объект
            # Записываем результаты
            for row in dataframe_to_rows(group_df, index=False, header=True):
                wb['Sheet'].append(row)

            # Форматирование итоговой таблицы
            # Ширина колонок
            wb['Sheet'].column_dimensions['A'].width = 15
            wb['Sheet'].column_dimensions['B'].width = 20
            wb['Sheet'].column_dimensions['C'].width = 36
            wb['Sheet'].column_dimensions['F'].width = 15
            wb['Sheet'].column_dimensions['G'].width = 15
            wb['Sheet'].column_dimensions['H'].width = 20
            # Перенос строк для заголовков
            wb['Sheet']['D1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['E1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['F1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['G1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['H1'].alignment = Alignment(wrap_text=True)

            wb.save(
                f'{path_to_end_folder}/Бурятия Контроль используемых площадей лесотаксационных выделов от {current_time}.xlsx')


        # Якутия
        elif region == 1:
            # СОздаем проверочный файл для проверки правильности ввода плошади выдела
            check_df = df.copy(deep=True)

            # Меняем тип столбца на строку чтобы создать строку включающую в себя все значения разделенные ;
            check_df['Площадь лесотаксационного выдела, га'] = check_df['Площадь лесотаксационного выдела, га'].astype(
                str)

            # Заполняем год лесоустройства на случай если он не заполнен или заполнен пробелами
            df['Год лесоустройства'] = df['Год лесоустройства'].astype(str)
            df['Год лесоустройства'] = df['Год лесоустройства'].apply(lambda x: x.replace(' ', 'Отсутствует'))
            df['Год лесоустройства'] = df['Год лесоустройства'].apply(lambda x: x.replace('nan', 'Отсутствует'))

            checked_pl = check_df.groupby(['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                           'Номер лесотаксационного выдела', 'Год лесоустройства']).agg(
                {'Площадь лесотаксационного выдела, га': combine})

            # Применяем функцию првоеряющую количество уникальных значений в столбце, если больше одного то значит есть ошибка в данных
            checked_pl['Контроль совпадения площади выдела'] = checked_pl['Площадь лесотаксационного выдела, га'].apply(
                check_unique)

            # переименовывам колонку
            checked_pl.rename(
                columns={'Площадь лесотаксационного выдела, га': 'Все значения площади для указанного выдела'},
                inplace=True)
            # Извлекаем индексы в колонки
            checked_pl = checked_pl.reset_index()

            # конвертируем в инт чтобы корректно отображалось
            checked_pl['Номер лесного квартала'] = checked_pl['Номер лесного квартала'].apply(prepare_column_to_int)
            checked_pl['Номер лесотаксационного выдела'] = checked_pl['Номер лесотаксационного выдела'].apply(
                prepare_column_to_int)
            checked_pl['Год лесоустройства'] = checked_pl['Год лесоустройства'].apply(prepare_column_to_int)

            # Сортируем
            checked_pl.sort_values(by=['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                       'Номер лесотаксационного выдела', 'Год лесоустройства'], inplace=True)

            # Получаем текущую дату
            current_time = time.strftime('%H_%M_%S %d.%m.%Y')
            # Сохраняем отчет
            # Для того чтобы увеличить ширину колонок для удобства чтения используем openpyxl
            wb = openpyxl.Workbook()  # Создаем объект
            # Записываем результаты
            for row in dataframe_to_rows(checked_pl, index=False, header=True):
                wb['Sheet'].append(row)

            # Форматирование итоговой таблицы
            # Ширина колонок
            wb['Sheet'].column_dimensions['A'].width = 15
            wb['Sheet'].column_dimensions['B'].width = 20
            wb['Sheet'].column_dimensions['C'].width = 36
            wb['Sheet'].column_dimensions['F'].width = 15
            wb['Sheet'].column_dimensions['G'].width = 15
            # Перенос строк для заголовков
            wb['Sheet']['D1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['E1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['F1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['G1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['H1'].alignment = Alignment(wrap_text=True)

            wb.save(
                f'{path_to_end_folder}/Саха(Якутия) Проверка правильности ввода площадей лесотаксационного выдела {current_time}.xlsx')

            # Основной отчет
            # Готовим колонки к группировке
            df['Площадь лесотаксационного выдела, га'] = df['Площадь лесотаксационного выдела, га'].astype(str)

            df['Площадь лесотаксационного выдела, га'] = df['Площадь лесотаксационного выдела, га'].apply(
                lambda x: x.replace(',', '.'))

            df['Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'] = df[
                'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'].astype(str)

            df['Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'] = df[
                'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'].apply(
                lambda x: x.replace(',', '.'))

            df['Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'] = df[
                'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га'].apply(
                convert_to_float)

            # Группируем
            group_df = df.groupby(['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                   'Номер лесотаксационного выдела', 'Год лесоустройства']).agg(
                {'Площадь лесотаксационного выдела, га': main_check_unique,
                 'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га': 'sum'})

            group_df = group_df.astype(str)

            # переименовываем колонку
            group_df.rename(columns={
                'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га': 'Используемая площадь лесотаксационного выдела, га'},
                inplace=True)

            # Извлекаем индексы в колонки
            group_df = group_df.reset_index()

            group_df['Площадь лесотаксационного выдела, га'] = group_df['Площадь лесотаксационного выдела, га'].apply(
                convert_to_float)
            group_df['Используемая площадь лесотаксационного выдела, га'] = group_df[
                'Используемая площадь лесотаксационного выдела, га'].apply(convert_to_float)

            group_df['Площадь лесотаксационного выдела, га'] = group_df['Площадь лесотаксационного выдела, га'].apply(
                convert_to_float)
            group_df['Используемая площадь лесотаксационного выдела, га'] = group_df[
                'Используемая площадь лесотаксационного выдела, га'].apply(convert_to_float)

            # Округляем до 3 знаков для корректного сравнения
            group_df['Площадь лесотаксационного выдела, га'] = np.round(
                group_df['Площадь лесотаксационного выдела, га'], decimals=3)
            group_df['Используемая площадь лесотаксационного выдела, га'] = np.round(
                group_df['Используемая площадь лесотаксационного выдела, га'], decimals=3)

            # Создаем колонку для контроля
            group_df['Контроль площади используемого надела'] = group_df['Площадь лесотаксационного выдела, га'] < \
                                                                group_df[
                                                                    'Используемая площадь лесотаксационного выдела, га']

            group_df['Контроль площади используемого надела'] = group_df['Контроль площади используемого надела'].apply(
                lambda x: 'Превышение используемой площади выдела!!!' if x is True else 'Все в порядке')

            # Изменяем состояние колонки если площадь всего выдела равна 0
            group_df['Контроль правильности ввода площади лесотаксационного выдела'] = group_df[
                'Площадь лесотаксационного выдела, га'].apply(
                lambda
                    x: 'Площадь лесотаксационного выдела равна нулю или  обнаружены разные значения площади выдела !!!' if x == 0 else 'Площади лесотаксационного выдела не отличаются друг от друга')

            # конвертируем в инт чтобы корректно отображалось
            group_df['Номер лесного квартала'] = group_df['Номер лесного квартала'].apply(prepare_column_to_int)
            group_df['Номер лесотаксационного выдела'] = group_df['Номер лесотаксационного выдела'].apply(
                prepare_column_to_int)
            group_df['Год лесоустройства'] = group_df['Год лесоустройства'].apply(prepare_column_to_int)

            # Сортируем
            group_df.sort_values(by=['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                     'Номер лесотаксационного выдела', 'Год лесоустройства'], inplace=True)

            # Сохраняем отчет
            # Для того чтобы увеличить ширину колонок для удобства чтения используем openpyxl
            wb = openpyxl.Workbook()  # Создаем объект
            # Записываем результаты
            for row in dataframe_to_rows(group_df, index=False, header=True):
                wb['Sheet'].append(row)

            # Форматирование итоговой таблицы
            # Ширина колонок
            wb['Sheet'].column_dimensions['A'].width = 15
            wb['Sheet'].column_dimensions['B'].width = 20
            wb['Sheet'].column_dimensions['C'].width = 36
            wb['Sheet'].column_dimensions['F'].width = 15
            wb['Sheet'].column_dimensions['G'].width = 15
            wb['Sheet'].column_dimensions['H'].width = 20
            # Перенос строк для заголовков
            wb['Sheet']['D1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['E1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['F1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['G1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['H1'].alignment = Alignment(wrap_text=True)

            wb.save(
                f'{path_to_end_folder}/Саха(Якутия) Контроль используемых площадей лесотаксационных выделов от {current_time}.xlsx')








    except ValueError as e:
        messagebox.showerror('Артемида 1.8',
                             f'Не найдена колонка или лист {e.args}\nДанные в файле должны находиться на листе с названием Реестр УПП'
                             f'\nКолонки 1-8 должны иметь названия: Лесничество,Участковое лесничество,Урочище ,Номер лесного квартала,\n'
                             f'Номер лесотаксационного выдела,Площадь лесотаксационного выдела, га,Обозначение части лесотаксационного выдела (лесопатологического выдела), га ,'
                             f'Площадь лесотаксационного выдела или его части (лесопатологического выдела), га')
    except NameError:
        messagebox.showerror('Артемида 1.8', f'Выберите файл с данными и конечную папку')
    except FileNotFoundError:
        messagebox.showerror('Артемида 1.8', f'Проверьте наличие указанных файлов')
    except PermissionError:
        messagebox.showerror('Артемида 1.8', f'Закройте файлы с созданными раньше отчетами!!!')
    except MemoryError:
        messagebox.showerror('Артемида 1.8', f'Слишком большая таблица!!!\nПроверьте размер таблицы!!!\nНажмите CTRL+END для проверки'
                                             f'Пересоздайте файл без пустых строк и столбцов')
    else:
        messagebox.showinfo('Артемида 1.8', 'Работа программы успешно завершена!!!')

def proccessing_report_purpose_category():
    """
    Функция для обработки реестра УПП, находит некорректно заполненные графы целевого назначения
    лесов и категории(графы 12 и графы 13)
    """
    try:
        region = group_rb_region_purpose_category.get()
        use_cols = list(range(25))
        df = pd.read_excel(reest_upp_purpose, skiprows=8,
                           usecols=use_cols)  # считываем датафрейм пропуская первые 8 строк,и загружая 25 строк

        """
        Соответствие названий колонок используемым в программе номерам колонок
        Лесничество -1
        Участковое лесничество- 2
        Урочище - 3
        Номер лесного квартала -4
        Номер лесотаксационного выдела -5
        Год лесоустройства -11
        Целевое назначение лесов - 12
        Категория защитных лесов (код) - 13
        """
        # Приводим названия колонок к строковому виду, чтобы избежать возможных проблем с названиями колонок
        df.columns = list(map(str, list(df.columns)))
        # на всякий случай очищаем от пробельных символов
        df.columns = list(map(lambda x: x.replace(" ", ""), list(df.columns)))

        df.rename(
            columns={'1': 'Лесничество', '2': 'Участковое лесничество', '3': 'Урочище', '4': 'Номер лесного квартала',
                     '5': 'Номер лесотаксационного выдела',
                     '11': 'Год лесоустройства', '12': 'Целевое назначение лесов',
                     '13': 'Категория защитных лесов (код)', }, inplace=True)

        # заполняем пропущенные места
        clean_column(df, ['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                          'Номер лесотаксационного выдела', 'Год лесоустройства'])

        # Меняем тип столбца на строку чтобы создать строку включающую в себя все значения разделенные ;заменяем нан на нули и очищаем от пробельных символов
        prepare_column_purpose_category(df, 'Целевое назначение лесов')
        prepare_column_purpose_category(df, 'Категория защитных лесов (код)')

        # Бурятия
        if region == 0:
            checked_pl = df.groupby(['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                     'Номер лесотаксационного выдела']).agg(
                {'Целевое назначение лесов': combine, 'Категория защитных лесов (код)': combine})

            # Извлекаем индекс
            out_df = checked_pl.reset_index()

            # Применяем функцию проверяющую количество уникальных значений в столбце, если больше одного то значит есть ошибка в данных
            out_df['Контроль правильности заполнения целевого назначения лесов'] = out_df[
                'Целевое назначение лесов'].apply(
                check_unique)
            out_df['Контроль правильности заполнения категории защитных лесов'] = out_df[
                'Категория защитных лесов (код)'].apply(
                check_unique)

            out_df['Контроль назначения лесов'] = out_df['Целевое назначение лесов'].apply(clean_purpose_column)

            out_df['Контроль назначения лесов'] = out_df['Контроль назначения лесов'].astype(
                int)  # Приводим на всякий случай к инту

            out_df['Контроль категории защитных лесов'] = out_df['Категория защитных лесов (код)'].apply(
                clean_purpose_column)
            out_df['Контроль категории защитных лесов'] = out_df['Контроль категории защитных лесов'].astype(
                int)  # Приводим на всякий случай к инту

            out_df.rename(columns={'Целевое назначение лесов': 'Показатели целевого назначения для данного выдела',
                                   'Категория защитных лесов (код)': 'Показатели категории защитных лесов для данного выдела'},
                          inplace=True)

            # находим итог
            out_df['Итоговый контроль защитных лесов'] = ((out_df['Контроль назначения лесов'] == 1) | (
                        out_df['Контроль назначения лесов'] == 0)) & (out_df['Контроль категории защитных лесов'] == 0)

            out_df['Итоговый контроль защитных лесов'] = out_df['Итоговый контроль защитных лесов'].apply(
                lambda
                    x: 'Ошибка, проверьте целевое назначение или категорию защитных лесов' if x == True else 'Все в порядке')

            # конвертируем в инт чтобы корректно отображалось
            out_df['Номер лесного квартала'] = out_df['Номер лесного квартала'].apply(prepare_column_to_int)
            out_df['Номер лесотаксационного выдела'] = out_df['Номер лесотаксационного выдела'].apply(
                prepare_column_to_int)

            # Сортируем
            out_df.sort_values(by=['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                   'Номер лесотаксационного выдела'], inplace=True)

            # Получаем текущую дату
            current_time = time.strftime('%H_%M_%S %d.%m.%Y')
            # Сохраняем отчет
            # Для того чтобы увеличить ширину колонок для удобства чтения используем openpyxl
            wb = openpyxl.Workbook()  # Создаем объект
            # Записываем результаты
            for row in dataframe_to_rows(out_df, index=False, header=True):
                wb['Sheet'].append(row)

            # Форматирование итоговой таблицы
            # Ширина колонок
            wb['Sheet'].column_dimensions['A'].width = 15
            wb['Sheet'].column_dimensions['B'].width = 20
            wb['Sheet'].column_dimensions['C'].width = 36
            wb['Sheet'].column_dimensions['F'].width = 15
            wb['Sheet'].column_dimensions['G'].width = 15
            wb['Sheet'].column_dimensions['H'].width = 15
            wb['Sheet'].column_dimensions['I'].width = 15
            wb['Sheet'].column_dimensions['J'].width = 15
            wb['Sheet'].column_dimensions['K'].width = 15
            wb['Sheet'].column_dimensions['L'].width = 15
            # Перенос строк для заголовков
            wb['Sheet']['F1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['G1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['H1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['I1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['L1'].alignment = Alignment(wrap_text=True)

            wb.save(
                f'{path_to_end_folder}/Бурятия Проверка правильности ввода целевого назначения лесов и категории защитных лесов {current_time}.xlsx')
        # Якутия
        elif region == 1:
            # Заполняем год лесоустройства на случай если он не заполнен или заполнен пробелами
            df['Год лесоустройства'] = df['Год лесоустройства'].astype(str)
            df['Год лесоустройства'] = df['Год лесоустройства'].apply(lambda x: x.replace(' ', 'Отсутствует'))
            df['Год лесоустройства'] = df['Год лесоустройства'].apply(lambda x: x.replace('nan', 'Отсутствует'))

            # Группируем
            checked_pl = df.groupby(['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                     'Номер лесотаксационного выдела', 'Год лесоустройства']).agg(
                {'Целевое назначение лесов': combine, 'Категория защитных лесов (код)': combine})
            # Извлекаем индекс
            out_df = checked_pl.reset_index()

            # Применяем функцию проверяющую количество уникальных значений в столбце, если больше одного то значит есть ошибка в данных
            out_df['Контроль правильности заполнения целевого назначения лесов'] = out_df[
                'Целевое назначение лесов'].apply(
                check_unique)
            out_df['Контроль правильности заполнения категории защитных лесов'] = out_df[
                'Категория защитных лесов (код)'].apply(
                check_unique)

            out_df['Контроль назначения лесов'] = out_df['Целевое назначение лесов'].apply(clean_purpose_column)

            out_df['Контроль назначения лесов'] = out_df['Контроль назначения лесов'].astype(
                int)  # Приводим на всякий случай к инту

            out_df['Контроль категории защитных лесов'] = out_df['Категория защитных лесов (код)'].apply(
                clean_purpose_column)
            out_df['Контроль категории защитных лесов'] = out_df['Контроль категории защитных лесов'].astype(
                int)  # Приводим на всякий случай к инту

            out_df.rename(columns={'Целевое назначение лесов': 'Показатели целевого назначения для данного выдела',
                                   'Категория защитных лесов (код)': 'Показатели категории защитных лесов для данного выдела'},
                          inplace=True)

            # находим итог
            out_df['Итоговый контроль защитных лесов'] = ((out_df['Контроль назначения лесов'] == 1) | (
                        out_df['Контроль назначения лесов'] == 0)) & (out_df['Контроль категории защитных лесов'] == 0)

            out_df['Итоговый контроль защитных лесов'] = out_df['Итоговый контроль защитных лесов'].apply(
                lambda
                    x: 'Ошибка, проверьте целевое назначение или категорию защитных лесов' if x == True else 'Все в порядке')

            # конвертируем в инт чтобы корректно отображалось
            out_df['Номер лесного квартала'] = out_df['Номер лесного квартала'].apply(prepare_column_to_int)
            out_df['Номер лесотаксационного выдела'] = out_df['Номер лесотаксационного выдела'].apply(
                prepare_column_to_int)
            out_df['Год лесоустройства'] = out_df['Год лесоустройства'].apply(prepare_column_to_int)

            # Сортируем
            out_df.sort_values(by=['Лесничество', 'Участковое лесничество', 'Урочище', 'Номер лесного квартала',
                                   'Номер лесотаксационного выдела', 'Год лесоустройства'], inplace=True)

            # Получаем текущую дату
            current_time = time.strftime('%H_%M_%S %d.%m.%Y')
            # Сохраняем отчет
            # Для того чтобы увеличить ширину колонок для удобства чтения используем openpyxl
            wb = openpyxl.Workbook()  # Создаем объект
            # Записываем результаты
            for row in dataframe_to_rows(out_df, index=False, header=True):
                wb['Sheet'].append(row)

            # Форматирование итоговой таблицы
            # Ширина колонок
            wb['Sheet'].column_dimensions['A'].width = 15
            wb['Sheet'].column_dimensions['B'].width = 20
            wb['Sheet'].column_dimensions['C'].width = 36
            wb['Sheet'].column_dimensions['F'].width = 15
            wb['Sheet'].column_dimensions['G'].width = 15
            wb['Sheet'].column_dimensions['H'].width = 15
            wb['Sheet'].column_dimensions['I'].width = 15
            wb['Sheet'].column_dimensions['J'].width = 15
            wb['Sheet'].column_dimensions['K'].width = 15
            wb['Sheet'].column_dimensions['L'].width = 15
            # Перенос строк для заголовков
            wb['Sheet']['F1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['G1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['H1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['I1'].alignment = Alignment(wrap_text=True)
            wb['Sheet']['L1'].alignment = Alignment(wrap_text=True)

            wb.save(
                f'{path_to_end_folder}/Саха(Якутия)Проверка правильности ввода целевого назначения лесов и категории защитных лесов {current_time}.xlsx')








    except ValueError as e:
        messagebox.showerror('Артемида 1.8',
                             f'Не найдена колонка или лист {e.args}\nДанные в файле должны находиться на листе с названием Реестр УПП'
                             f'\nВ файле должны быть колонки: Лесничество,Участковое лесничество,Урочище ,Номер лесного квартала,\n'
                             f'Номер лесотаксационного выдела,Целевое назначение лесов ,Категория защитных лесов (код) ')
    except NameError:
        messagebox.showerror('Артемида 1.8', f'Выберите файл с данными и конечную папку')
    except FileNotFoundError:
        messagebox.showerror('Артемида 1.8', f'Проверьте наличие указанных файлов')
    except PermissionError:
        messagebox.showerror('Артемида 1.8', f'Закройте файлы с созданными раньше отчетами!!!')
    except MemoryError:
        messagebox.showerror('Артемида 1.8', f'Слишком большая таблица!!!\nПроверьте размер таблицы!!!\nНажмите CTRL+END для проверки'
                                             f'Пересоздайте файл без пустых строк и столбцов')
    else:
        messagebox.showinfo('Артемида 1.8', 'Работа программы успешно завершена!!!')


def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except ValueError:
        return ''

def proccessing_transfer_table3_to_reestr():
    """
    Функция для переноса подходящих данных из таблицы 3 в реестр УПП
    :return:
    """
    try:

        # Получаем значение кнопки
        region = group_rb_region_transfer_3_to_reestr.get()
        # Создаем список колонок которые нужно загрузить
        use_cols = list(range(25))

        # Загружаем датафреймы
        df_upp = pd.read_excel(file_transfer_reestr, skiprows=8, usecols=use_cols, dtype={4: str, 5: str, 11: str})
        df_table_3 = pd.read_excel(file_transfer_to_upp, skiprows=6,
                                   usecols=[4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17, 18, 19, 21, 22, 23, 24, 25,
                                            26, 27, 28,
                                            29, 32], dtype={7: str, 8: str, 9: str, 15: str})

        # Приводим названия колонок к строковому виду, чтобы избежать возможных проблем с названиями колонок
        df_upp.columns = list(map(str, list(df_upp.columns)))
        df_table_3.columns = list(map(str, list(df_table_3.columns)))

        # Очищаем на всякий случай от пробельных символов
        df_upp.columns = list(map(lambda x: x.strip(), list(df_upp.columns)))
        df_table_3.columns = list(map(lambda x: x.strip(), list(df_table_3.columns)))

        # Фильтруем датафрейм отбирая только те записи в которых есть 1 в графе 30
        transfer_df = df_table_3[df_table_3['30'] == 1]

        # перемещаем площадь выдела
        transfer_df.insert(7, '12', transfer_df['33'])

        # удаляем лишний столбец с площадью выдела и признаков внесения в реестр
        transfer_df.drop(columns=['30', '33'], inplace=True)

        transfer_df['17'] = transfer_df['17'].astype(str)  # Приводим колонку к строковому формату

        transfer_df['17'] = transfer_df['17'].apply(lambda x: x.replace(' ', ''))
        # Заменяем категории таблицы 3 на категории Реестра УПП
        transfer_df['17'] = transfer_df['17'].replace(
            regex={'120': '2', '100': '1', r'130|131|132|133|134|135|136': '3',
                   r'141|142|143|144|145|146|147|148|149|150|151|152': '4', 'nan': '0'})

        transfer_df.columns = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16',
                               '17', '18', '19', '20', '21', '22', '23', '24']

        # Бурятия
        if region == 0:

            # Создаем датафрейм для проверки наличия записи в реестре
            checked_df = df_upp[['1', '2', '3', '4', '5']]
            checked_df = checked_df.astype(str)  # делаем данные строковыми

            # Готовим 4 и 5 колонки чтобы они были разделеныпри склеивании
            checked_df['4'] = checked_df['4'].apply(lambda x: 'кв.' + x)  # Добавляем разделитель квартал
            checked_df['5'] = checked_df['5'].apply(lambda x: 'в.' + x)  # Добавляем разделитель выдел

            # Создаем в каждом датафрейме колонку с айди путем склеивания всех нужных колонок в одну строку
            checked_df['ID'] = checked_df.loc[:, ['1', '2', '3', '4', '5']].sum(axis=1)
            # Удаляем пробелы
            checked_df['ID'] = checked_df['ID'].apply(lambda x: x.replace(' ', ''))
            checked_df['ID'] = checked_df['ID'].apply(lambda x: x.replace('nan', ''))

            # делаем строковыми первые 5 колонок
            transfer_df[['1', '2', '3', '4', '5', ]] = transfer_df[['1', '2', '3', '4', '5', ]].astype(str)

            # Готовим 4 и 5 колонки чтобы они были разделеныпри склеивании
            transfer_df['4'] = transfer_df['4'].apply(lambda x: 'кв.' + x)  # Добавляем разделитель квартал
            transfer_df['5'] = transfer_df['5'].apply(lambda x: 'в.' + x)  # Добавляем разделитель выдел

            transfer_df['ID'] = transfer_df.loc[:, ['1', '2', '3', '4', '5']].sum(axis=1)
            transfer_df['ID'] = transfer_df['ID'].apply(lambda x: x.replace(' ', ''))
            transfer_df['ID'] = transfer_df['ID'].apply(lambda x: x.replace('nan', ''))

            # Мержим по полю айди
            merged_df = pd.merge(checked_df, transfer_df, how='outer', left_on='ID', right_on='ID', indicator=True)

            # Отбираем только те значения которые есть в правом датафрейме
            added_df = merged_df[merged_df['_merge'] == 'right_only']

            added_df.drop(columns=['1_x', '2_x', '3_x', '4_x', '5_x', 'ID', '_merge'],
                          inplace=True)  # удаляем лишние колонки

            added_df.rename(columns={'1_y': '1', '2_y': '2', '3_y': '3', '4_y': '4', '5_y': '5'},
                            inplace=True)  # переименовываем колонки для корректного добавления

            itog_df = pd.concat([df_upp, added_df], ignore_index=True)

            # Приводим даты к нормальному виду ДД.ММ.ГГГГ
            itog_df['17'] = pd.to_datetime(itog_df['17'], errors='coerce', dayfirst=True)
            itog_df['22'] = pd.to_datetime(itog_df['22'], errors='coerce', dayfirst=True)
            itog_df['17'] = itog_df['17'].apply(create_doc_convert_date)
            itog_df['22'] = itog_df['22'].apply(create_doc_convert_date)

            # Заменяем nan и пробелы в итоговом файле
            # приводим к строковому виду
            itog_df[['1', '2', '3']] = itog_df[['1', '2', '3']].astype(str)
            # Заменяем
            itog_df[['1', '2', '3']] = itog_df[['1', '2', '3']].apply(lambda x: x.replace(' ', ''))

            itog_df[['1', '2', '3', '4', '5']] = itog_df[['1', '2', '3', '4', '5']].apply(
                lambda x: x.replace('nan', np.nan))

            # Приводим к строковому виду чтобы очистить от разделителей кв. и в.
            itog_df['4'] = itog_df['4'].astype(str)
            itog_df['5'] = itog_df['5'].astype(str)

            # Очищаем от разделителей
            itog_df['4'] = itog_df['4'].apply(lambda x: x.replace('кв.', ''))
            itog_df['5'] = itog_df['5'].apply(lambda x: x.replace('в.', ''))

            # конвертируем в инт чтобы корректно отображалось
            itog_df['4'] = itog_df['4'].apply(prepare_column_to_int)
            itog_df['5'] = itog_df['5'].apply(prepare_column_to_int)

            # Сортируем
            itog_df.sort_values(by=['1', '2', '3', '4', '5'], inplace=True)

            # Получаем текущую дату
            current_time = time.strftime('%H_%M_%S %d.%m.%Y')
            # Сохраняем отчет
            # Для того чтобы увеличить ширину колонок для удобства чтения используем openpyxl
            wb = load_workbook(header_reestr)  # Создаем объект
            # Записываем результаты
            for row in dataframe_to_rows(itog_df, index=False, header=False):
                wb['Реестр УПП'].append(row)

            # Заменяем nan и пробелы в таблице с добавленными участками
            # приводим к строковому виду
            added_df[['1', '2', '3']] = added_df[['1', '2', '3']].astype(str)
            added_df[['1', '2', '3']] = added_df[['1', '2', '3']].apply(lambda x: x.replace(' ', ''))
            added_df[['1', '2', '3']] = added_df[['1', '2', '3']].apply(lambda x: x.replace('nan', ''))

            # Приводим к строковому виду чтобы очистить от разделителей кв. и в.
            added_df['4'] = added_df['4'].astype(str)
            added_df['5'] = added_df['5'].astype(str)

            # Очищаем от разделителей
            added_df['4'] = added_df['4'].apply(lambda x: x.replace('кв.', ''))
            added_df['5'] = added_df['5'].apply(lambda x: x.replace('в.', ''))

            # конвертируем в инт чтобы корректно отображалось
            added_df['4'] = added_df['4'].apply(prepare_column_to_int)
            added_df['5'] = added_df['5'].apply(prepare_column_to_int)

            # Переименовываем после соединения колонки в таблице с добавленными участками
            added_df.rename(columns={'1': 'Лесничество', '2': 'Участковое лесничество', '3': 'Урочище',
                                     '4': 'Номер лесного квартала', '5': 'Номер лесотаксационного выдела',
                                     }, inplace=True)

            # Сохраняем файл с добавляемыми данными, чтобы пользователи знали что именно добавилось
            added_df.to_excel(
                f'{path_to_end_folder}/Бурятия Участки из таблицы 3 добавленные в реестр УПП от {current_time}.xlsx',
                index=False)

            wb.save(f'{path_to_end_folder}/Бурятия Реестр УПП с добавлением данных из таблицы 3 {current_time}.xlsx')

            # Якутия
        elif region == 1:
            # Создаем датафрейм для проверки наличия записи в реестре
            checked_df = df_upp[['1', '2', '3', '4', '5', '11']]
            checked_df = checked_df.astype(str)  # делаем данные строковыми

            checked_df['4'] = checked_df['4'].apply(lambda x: 'кв.' + x)  # Добавляем разделитель квартал
            checked_df['5'] = checked_df['5'].apply(lambda x: 'в.' + x)  # Добавляем разделитель выдел
            checked_df['11'] = checked_df['11'].apply(lambda x: 'г.' + x)  # Добавляем разделитель год

            # Создаем в каждом датафрейме колонку с айди путем склеивания всех нужных колонок в одну строку
            checked_df['ID'] = checked_df.loc[:, ['1', '2', '3', '4', '5', '11']].sum(axis=1)

            # Удаляем пробелы
            checked_df['ID'] = checked_df['ID'].apply(lambda x: x.replace(' ', ''))
            checked_df['ID'] = checked_df['ID'].apply(lambda x: x.replace('nan', ''))

            # делаем строковыми первые 6 колонок в таблице 3
            transfer_df[['1', '2', '3', '4', '5', '11']] = transfer_df[['1', '2', '3', '4', '5', '11']].astype(str)

            # Готовим 4 и 5 11 колонки чтобы они были разделеныпри склеивании
            transfer_df['4'] = transfer_df['4'].apply(lambda x: 'кв.' + x)  # Добавляем разделитель квартал
            transfer_df['5'] = transfer_df['5'].apply(lambda x: 'в.' + x)  # Добавляем разделитель выдел
            transfer_df['11'] = transfer_df['11'].apply(lambda x: 'г.' + x)  # Добавляем разделитель год

            transfer_df['ID'] = transfer_df.loc[:, ['1', '2', '3', '4', '5', '11']].sum(axis=1)
            transfer_df['ID'] = transfer_df['ID'].apply(lambda x: x.replace(' ', ''))
            transfer_df['ID'] = transfer_df['ID'].apply(lambda x: x.replace('nan', ''))

            # Мержим по полю айди
            merged_df = pd.merge(checked_df, transfer_df, how='outer', left_on='ID', right_on='ID', indicator=True)

            # Отбираем только те значения которые есть в правом датафрейме
            added_df = merged_df[merged_df['_merge'] == 'right_only']

            added_df.drop(columns=['1_x', '2_x', '3_x', '4_x', '5_x', '11_x', 'ID', '_merge'],
                          inplace=True)  # удаляем лишние колонки

            added_df.rename(columns={'1_y': '1', '2_y': '2', '3_y': '3', '4_y': '4', '5_y': '5', '11_y': '11'},
                            inplace=True)  # переименовываем колонки для корректно

            itog_df = pd.concat([df_upp, added_df], ignore_index=True)

            # Приводим даты к нормальному виду ДД.ММ.ГГГГ
            itog_df['17'] = pd.to_datetime(itog_df['17'], errors='coerce', dayfirst=True)
            itog_df['22'] = pd.to_datetime(itog_df['22'], errors='coerce', dayfirst=True)
            itog_df['17'] = itog_df['17'].apply(create_doc_convert_date)
            itog_df['22'] = itog_df['22'].apply(create_doc_convert_date)

            # Заменяем nan и пробелы в итоговом файле
            # приводим к строковому виду
            itog_df[['1', '2', '3']] = itog_df[['1', '2', '3']].astype(str)
            # Заменяем
            itog_df[['1', '2', '3']] = itog_df[['1', '2', '3']].apply(lambda x: x.replace(' ', ''))
            itog_df[['1', '2', '3', '4', '5', '11']] = itog_df[['1', '2', '3', '4', '5', '11']].apply(
                lambda x: x.replace('nan', np.nan))

            # Приводим к строковому виду чтобы очистить от разделителей кв. и в.
            itog_df['4'] = itog_df['4'].astype(str)
            itog_df['5'] = itog_df['5'].astype(str)
            itog_df['11'] = itog_df['11'].astype(str)

            # Очищаем от разделителей
            itog_df['4'] = itog_df['4'].apply(lambda x: x.replace('кв.', ''))
            itog_df['5'] = itog_df['5'].apply(lambda x: x.replace('в.', ''))
            itog_df['11'] = itog_df['11'].apply(lambda x: x.replace('г.', ''))

            # конвертируем в инт чтобы корректно отображалось
            itog_df['4'] = itog_df['4'].apply(prepare_column_to_int)
            itog_df['5'] = itog_df['5'].apply(prepare_column_to_int)
            itog_df['11'] = itog_df['11'].apply(prepare_column_to_int)

            # Сортируем
            itog_df.sort_values(by=['1', '2', '3', '4', '5', '11'], inplace=True)

            # Получаем текущую дату
            current_time = time.strftime('%H_%M_%S %d.%m.%Y')
            # Сохраняем отчет
            # Для того чтобы увеличить ширину колонок для удобства чтения используем openpyxl
            wb = load_workbook(header_reestr)  # Создаем объект
            # Записываем результаты
            for row in dataframe_to_rows(itog_df, index=False, header=False):
                wb['Реестр УПП'].append(row)

            # Заменяем nan и пробелы в таблице с добавленными участками
            # приводим к строковому виду
            added_df[['1', '2', '3', '11']] = added_df[['1', '2', '3', '11']].astype(str)
            added_df[['1', '2', '3', '11']] = added_df[['1', '2', '3', '11']].apply(lambda x: x.replace(' ', ''))
            added_df[['1', '2', '3', '11']] = added_df[['1', '2', '3', '11']].apply(lambda x: x.replace('nan', ''))

            #   Приводим к строковому виду чтобы очистить от разделителей кв. и в.
            added_df['4'] = added_df['4'].astype(str)
            added_df['5'] = added_df['5'].astype(str)
            added_df['11'] = added_df['11'].astype(str)

            # Очищаем от разделителей
            added_df['4'] = added_df['4'].apply(lambda x: x.replace('кв.', ''))
            added_df['5'] = added_df['5'].apply(lambda x: x.replace('в.', ''))
            added_df['11'] = added_df['11'].apply(lambda x: x.replace('г.', ''))

            # конвертируем в инт год лесоустройства
            # конвертируем в инт чтобы корректно отображалось
            added_df['4'] = added_df['4'].apply(prepare_column_to_int)
            added_df['5'] = added_df['5'].apply(prepare_column_to_int)
            added_df['11'] = added_df['11'].apply(prepare_column_to_int)

            # Сортируем
            added_df.sort_values(by=['1', '2', '3', '4', '5', '11'], inplace=True)

            # Переименовываем после соединения колонки в таблице с добавленными участками
            added_df.rename(columns={'1': 'Лесничество', '2': 'Участковое лесничество', '3': 'Урочище',
                                     '4': 'Номер лесного квартала', '5': 'Номер лесотаксационного выдела', },
                            inplace=True)

            # Сохраняем файл с добавляемыми данными, чтобы пользователи знали что именно добавилось
            added_df.to_excel(
                f'{path_to_end_folder}/Саха(Якутия) Участки из таблицы 3 добавленные в реестр УПП от {current_time}.xlsx',
                index=False)

            wb.save(
                f'{path_to_end_folder}/Саха(Якутия) Реестр УПП с добавлением данных из таблицы 3 {current_time}.xlsx')

    except ValueError as e:
        messagebox.showerror('Артемида 1.8',
                             f'Не найдена колонка или лист {e.args}\nДанные в файле должны находиться на листе с названием Реестр УПП'
                             f'\nВ файле должны быть колонки: Лесничество,Участковое лесничество,Урочище ,Номер лесного квартала,\n'
                             f'Номер лесотаксационного выдела,Целевое назначение лесов ,Категория защитных лесов (код) ')
    except NameError:
        messagebox.showerror('Артемида 1.8', f'Выберите файл с данными и конечную папку')
    except FileNotFoundError:
        messagebox.showerror('Артемида 1.8', f'Проверьте наличие указанных файлов')
    except PermissionError:
        messagebox.showerror('Артемида 1.8', f'Закройте файлы с созданными раньше отчетами!!!')
    except MemoryError:
        messagebox.showerror('Артемида 1.8', f'Слишком большая таблица!!!\nПроверьте размер таблицы!!!\nНажмите CTRL+END для проверки'
                                             f'Пересоздайте файл без пустых строк и столбцов')
    else:
        messagebox.showinfo('Артемида 1.8', 'Работа программы успешно завершена!!!')


if __name__ == '__main__':
    window = Tk()
    window.title('Артемида 1.8')
    window.geometry('760x750+600+200')
    window.resizable(False, False)

    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)


    # Создаем вкладку обработки данных по площадям выделов
    tab_report_square = ttk.Frame(tab_control)
    tab_control.add(tab_report_square, text='Контроль площадей\nвыделов')
    tab_control.pack(expand=1, fill='both')

    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_report_square,
                      text='Центр защиты леса Республики Бурятия\n'
                           'Контроль соответствия площадей выделов,\nправильности заполнения площадей выделов ')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')

    img = PhotoImage(file=path_to_img)
    Label(tab_report_square,
          image=img
          ).grid(column=1, row=0, padx=10, pady=25)

    # Переключатель:Бурятия или Якутия
    # Создаем переменную хранящую регион, в зависимости от значения будет происходить обработка
    group_rb_region_report_square = IntVar()
    group_rb_region_report_square.set(0) # Устанавливаем значение по умолчанию
    # Создаем фрейм для размещения переключателей(pack и грид не используются в одном контейнере)
    frame_rb_region_report_square = LabelFrame(tab_report_square, text='1) Выберите регион')
    frame_rb_region_report_square.grid(column=0, row=1, padx=10)
    Radiobutton(frame_rb_region_report_square, text='Бурятия', variable=group_rb_region_report_square, value=0).pack()
    Radiobutton(frame_rb_region_report_square, text='Саха(Якутия)', variable=group_rb_region_report_square, value=1).pack()




    # Создаем кнопку Выбрать файл с данными
    btn_choose_data = Button(tab_report_square, text='2) Выберите файл\nреестра УПП', font=('Arial Bold', 20),
                             command=select_file_data_xlsx
                             )
    btn_choose_data.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder = Button(tab_report_square, text='3) Выберите конечную папку', font=('Arial Bold', 20),
                                   command=select_end_folder
                                   )
    btn_choose_end_folder.grid(column=0, row=4, padx=10, pady=10)

    # Создаем кнопку обработки данных

    btn_proccessing_data = Button(tab_report_square, text='4) Обработать данные', font=('Arial Bold', 20),
                                  command=processing_report_square_wood
                                  )
    btn_proccessing_data.grid(column=0, row=5, padx=10, pady=10)

    """
    Создаем вкладку обработки данных по целевому назначению и категории защищености
    """
    tab_report_purpose_category = ttk.Frame(tab_control)
    tab_control.add(tab_report_purpose_category, text='Контроль назначения и\n категории защитности')

    # Создаем метку для описания назначения программы
    lbl_hello_purpose = Label(tab_report_purpose_category,
                      text='Центр защиты леса Республики Бурятия\n'
                           'Контроль соответствия целевого назначения лесов\nи категории защитности')
    lbl_hello_purpose.grid(column=0, row=0, padx=10, pady=25)


    # Картинка
    path_to_img = resource_path('logo.png')

    img_purpose = PhotoImage(file=path_to_img)
    Label(tab_report_purpose_category,
          image=img_purpose
          ).grid(column=1, row=0, padx=10, pady=25)

    # Переключатель:Бурятия или Якутия
    # Создаем переменную хранящую регион, в зависимости от значения будет происходить обработка
    group_rb_region_purpose_category = IntVar()
    group_rb_region_purpose_category.set(0)  # Устанавливаем значение по умолчанию
    # Создаем фрейм для размещения переключателей(pack и грид не используются в одном контейнере)
    frame_rb_region_purpose_category = LabelFrame(tab_report_purpose_category, text='1) Выберите регион')
    frame_rb_region_purpose_category.grid(column=0, row=1, padx=10)
    Radiobutton(frame_rb_region_purpose_category, text='Бурятия', variable=group_rb_region_purpose_category, value=0).pack()
    Radiobutton(frame_rb_region_purpose_category, text='Саха(Якутия)', variable=group_rb_region_purpose_category, value=1).pack()

    # Создаем кнопку Выбрать файл с данными
    btn_choose_data_purpose = Button(tab_report_purpose_category, text='2) Выберите файл\nреестра УПП', font=('Arial Bold', 20),
                             command=select_file_reestr_purpose
                             )
    btn_choose_data_purpose.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder_purpose = Button(tab_report_purpose_category, text='3) Выберите конечную папку', font=('Arial Bold', 20),
                                   command=select_end_folder
                                   )
    btn_choose_end_folder_purpose.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку обработки данных

    btn_proccessing_data_purpose = Button(tab_report_purpose_category, text='4) Обработать данные', font=('Arial Bold', 20),
                                  command=proccessing_report_purpose_category
                                  )
    btn_proccessing_data_purpose.grid(column=0, row=4, padx=10, pady=10)


    """
    Создаем вкладку обработки данных по проверке наличия записи в реестре
    """
    tab_presense_reestr = ttk.Frame(tab_control)
    tab_control.add(tab_presense_reestr, text='Контроль наличия\nвыделов в УПП')

    # Создаем метку для описания назначения программы
    lbl_hello_presense = Label(tab_presense_reestr,
                      text='Центр защиты леса Республики Бурятия\n'
                           'Сравнение УПП с другими ведомостями\nна наличие участков\nили их отсутствие')
    lbl_hello_presense.grid(column=0, row=0, padx=10, pady=25)


    # Картинка
    path_to_img = resource_path('logo.png')

    img_presense = PhotoImage(file=path_to_img)
    Label(tab_presense_reestr,
          image=img_presense
          ).grid(column=1, row=0, padx=10, pady=25)

     # Создаем кнопку Выбрать файл с реестром
    btn_choose_reestr_presense = Button(tab_presense_reestr, text='1) Выберите файл\n реестра УПП', font=('Arial Bold', 20),
                             command=select_file_reestr_presense
                             )
    btn_choose_reestr_presense.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку Выбрать файл с ведомостью
    btn_choose_statement_presense = Button(tab_presense_reestr, text='2) Выберите файл ведомости', font=('Arial Bold', 20),
                             command=select_file_statement_presense
                             )
    btn_choose_statement_presense.grid(column=0, row=3, padx=10, pady=10)


    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder_presense = Button(tab_presense_reestr, text='3) Выберите конечную папку', font=('Arial Bold', 20),
                                   command=select_end_folder
                                   )
    btn_choose_end_folder_presense.grid(column=0, row=4, padx=10, pady=10)

    # Создаем кнопку обработки данных

    btn_proccessing_data_presense = Button(tab_presense_reestr, text='4) Обработать данные', font=('Arial Bold', 20),
                                  command=processing_presense_reestr
                                  )
    btn_proccessing_data_presense.grid(column=0, row=5, padx=10, pady=10)

    """
    Создаем вкладку Перенос данных из таблицы 3 в реестр
    """
    tab_transfer_3_to_reestr = ttk.Frame(tab_control)
    tab_control.add(tab_transfer_3_to_reestr, text='Внесение данных\nиз таблицы 3 в УПП')

    # Создаем метку для описания назначения программы
    lbl_hello_transfer_3_to_reestr = Label(tab_transfer_3_to_reestr,
                      text='Центр защиты леса Республики Бурятия\n'
                           'Внесение данных из таблицы 3\n в реестр УПП')
    lbl_hello_transfer_3_to_reestr.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')

    img_transfer_3_to_reestr = PhotoImage(file=path_to_img)
    Label(tab_transfer_3_to_reestr,
          image=img_transfer_3_to_reestr
          ).grid(column=1, row=0, padx=10, pady=25)

    # Переключатель:Бурятия или Якутия
    # Создаем переменную хранящую регион, в зависимости от значения будет происходить обработка
    group_rb_region_transfer_3_to_reestr = IntVar()
    group_rb_region_transfer_3_to_reestr.set(0)  # Устанавливаем значение по умолчанию
    # Создаем фрейм для размещения переключателей(pack и грид не используются в одном контейнере)
    frame_rb_region_transfer_3_to_reestr = LabelFrame(tab_transfer_3_to_reestr, text='1) Выберите регион')
    frame_rb_region_transfer_3_to_reestr.grid(column=0, row=1, padx=10)
    Radiobutton(frame_rb_region_transfer_3_to_reestr, text='Бурятия', variable=group_rb_region_transfer_3_to_reestr, value=0).pack()
    Radiobutton(frame_rb_region_transfer_3_to_reestr, text='Саха(Якутия)', variable=group_rb_region_transfer_3_to_reestr, value=1).pack()


    # Создаем кнопку Выбрать файл с заголовком
    btn_choose_header_transfer_3_to_reestr = Button(tab_transfer_3_to_reestr, text='2) Выберите файл\n с заголовком файла реестра УПП', font=('Arial Bold', 18),
                             command=select_header_reestr
                             )
    btn_choose_header_transfer_3_to_reestr.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку выбора файла с реестром
    btn_choose_reestr_transfer3 = Button(tab_transfer_3_to_reestr, text='3) Выберите файл\nреестра УПП', font=('Arial Bold', 18),
                             command=select_reestr_transfer3
                             )
    btn_choose_reestr_transfer3.grid(column=0, row=3, padx=10, pady=10)

    #Создаем кнопку выбора файла с таблицей 3
    btn_choose_table3_transfer3 = Button(tab_transfer_3_to_reestr, text='4) Выберите файл\nс таблицей 3', font=('Arial Bold', 18),
                             command=select_table3_transfer3
                             )
    btn_choose_table3_transfer3.grid(column=0, row=4, padx=10, pady=10)

    # Создаем кнопку выбора конечной папки
    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder_transfer3 = Button(tab_transfer_3_to_reestr, text='5) Выберите конечную папку', font=('Arial Bold', 18),
                                   command=select_end_folder
                                   )
    btn_choose_end_folder_transfer3.grid(column=0, row=5, padx=10, pady=10)

    # Создаем кнопку обработки данных

    btn_proccessing_transfer_table3_to_reestr = Button(tab_transfer_3_to_reestr, text='6) Обработать данные', font=('Arial Bold', 18),
                                  command=proccessing_transfer_table3_to_reestr
                                  )
    btn_proccessing_transfer_table3_to_reestr.grid(column=0, row=6, padx=10, pady=10)


    window.mainloop()
