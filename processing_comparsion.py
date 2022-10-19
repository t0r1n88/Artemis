import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import time
import datetime
from dateutil.parser import ParserError
import warnings
pd.options.mode.chained_assignment = None
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=FutureWarning)

def convert_params_columns_to_int(lst):
    """
    Функция для конвератации значений колонок которые нужно обработать.
    Очищает от пустых строк, чтобы в итоге остался список из чисел в формате int
    """
    out_lst = [] # Создаем список в который будем добавлять только числа
    for value in lst: # Перебираем список
        try:
            # Обрабатываем случай с нулем, для того чтобы после приведения к питоновскому отсчету от нуля не получилась колонка с номером -1
            number = int(value)
            if number != 0:
                out_lst.append(value) # Если конвертирования прошло без ошибок то добавляем
            else:
                continue
        except: # Иначе пропускаем
            continue
    return out_lst

def convert_columns_to_str(df, number_columns):
    """
    Функция для конвертации указанных столбцов в строковый тип и очистки от пробельных символов в начале и конце
    """

    for column in number_columns:  # Перебираем список нужных колонок
        try:
            df.iloc[:, column] = df.iloc[:, column].astype(str)
            # Очищаем колонку от пробельных символов с начала и конца
            df.iloc[:, column] = df.iloc[:, column].apply(lambda x: x.strip())
        except IndexError:
            print('index error')
            # messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.16',
            #                      'Проверьте порядковые номера колонок которые вы хотите обработать.')


def processing_date_column(df, lst_columns):
    """
    Функция для обработки столбцов с датами. конвертация в строку формата ДД.ММ.ГГГГ
    """
    # получаем первую строку
    first_row = df.iloc[0, lst_columns]

    lst_first_row = list(first_row)  # Превращаем строку в список
    lst_date_columns = []  # Создаем список куда будем сохранять колонки в которых находятся даты
    tupl_row = list(zip(lst_columns,
                        lst_first_row))  # Создаем список кортежей формата (номер колонки,значение строки в этой колонке)

    for idx, value in tupl_row:  # Перебираем кортеж
        result = check_date_columns(idx, value)  # проверяем является ли значение датой
        if result:  # если да то добавляем список порядковый номер колонки
            lst_date_columns.append(result)
        else:  # иначе проверяем следующее значение
            continue
    for i in lst_date_columns:  # Перебираем список с колонками дат, превращаем их в даты и конвертируем в нужный строковый формат
        df.iloc[:, i] = pd.to_datetime(df.iloc[:, i], errors='coerce', dayfirst=True)
        df.iloc[:, i] = df.iloc[:, i].apply(create_doc_convert_date)

def check_date_columns(i, value):
    """
    Функция для проверки типа колонки. Необходимо найти колонки с датой
    :param i:
    :param value:
    :return:
    Просто пытаемся сконвертировать в значение в дату , если все прошло успешно то возвращаем номер колонки где это удалось
    """
    try:
        itog = pd.to_datetime(str(value), infer_datetime_format=True)

    except ParserError:
        pass
    except ValueError:
        pass
    except TypeError:
        pass
    else:
        return i

def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except ValueError:
        return 'Не удалось конвертировать дату.Проверьте значение ячейки!!!'
    except TypeError:
        return 'Не удалось конвертировать дату.Проверьте значение ячейки!!!'


path_to_end_folder_comparison = 'data/'
# first_sheet_name = entry_first_sheet_name.get()
# second_sheet_name = entry_second_sheet_name.get()
first_sheet_name = 'Реестр УПП'
second_sheet_name = 'Zakamenskoe_III_2022'

skip_rows_first_df = 6
skip_rows_second_df = 0

name_first_file_comparison = 'data/Сравнение УПП с другими ведомостями на наличие участков или их отсутствие/2022-10-27_64_Реестр УПП с дополнительными колонками..xlsx'
name_second_file_comparison = 'data/Сравнение УПП с другими ведомостями на наличие участков или их отсутствие/Ведомость.xlsx'
file_params = 'data/params.xlsx'

# загружаем файлы
# На случай если
first_df = pd.read_excel(name_first_file_comparison, sheet_name=first_sheet_name,skiprows=skip_rows_first_df, dtype=str, keep_default_na=False)
second_df = pd.read_excel(name_second_file_comparison, sheet_name=second_sheet_name,skiprows=skip_rows_second_df, dtype=str, keep_default_na=False)
params = pd.read_excel(file_params, header=None, keep_default_na=False)

# Преврашаем каждую колонку в список
params_first_columns = params[0].tolist()
params_second_columns = params[1].tolist()

# Конвертируем в инт заодно проверяя корректность введенных данных
int_params_first_columns = convert_params_columns_to_int(params_first_columns)
int_params_second_columns = convert_params_columns_to_int(params_second_columns)

# Отнимаем 1 от каждого значения чтобы привести к питоновским индексам
int_params_first_columns = list(map(lambda x: x - 1, int_params_first_columns))
int_params_second_columns = list(map(lambda x: x - 1, int_params_second_columns))

# Конвертируем нужные нам колонки в str
convert_columns_to_str(first_df, int_params_first_columns)
convert_columns_to_str(second_df, int_params_second_columns)

# в этом месте конвертируем даты в формат ДД.ММ.ГГГГ
processing_date_column(first_df, int_params_first_columns)
processing_date_column(second_df, int_params_second_columns)

# Создаем в каждом датафрейме колонку с айди путем склеивания всех нужных колонок в одну строку
first_df['ID'] = first_df.iloc[:, int_params_first_columns].sum(axis=1)
second_df['ID'] = second_df.iloc[:, int_params_second_columns].sum(axis=1)

# очищаем от пробелов между словами
first_df['ID'] = first_df['ID'].apply(lambda x: x.replace(' ', ''))
second_df['ID'] = second_df['ID'].apply(lambda x: x.replace(' ', ''))

# Обрабатываем дубликаты

duplicates_first_df = first_df[first_df.duplicated(subset=['ID'],
                                                   keep=False)]  # Сохраняем все значения у которых есть дубликаты в отдельный датафрейм

first_df.drop_duplicates(subset=['ID'], keep=False, inplace=True)  # Удаляем дубликаты из датафрейма

duplicates_second_df = second_df[second_df.duplicated(subset=['ID'],
                                                      keep=False)]  # Сохраняем все значения у которых есть дубликаты в отдельный датафрейм
second_df.drop_duplicates(subset=['ID'], keep=False, inplace=True)  # Удаляем дубликаты из датафрейма

# Проверяем размер датафрейма с дубликатами, если он больше 0 то выдаем сообшение пользователю
# if duplicates_first_df.shape[0] > 0:
#     messagebox.showwarning('Веста Обработка таблиц и создание документов ver 1.16',
#                            f'В первой таблице обнаружены дубликаты!!!\nДля корректного объединения таблиц ,дубликаты перенесены в отдельный лист итоговой таблицы')
# if duplicates_second_df.shape[0] > 0:
#     messagebox.showwarning('Веста Обработка таблиц и создание документов ver 1.16',
#                            f'Во второй таблице обнаружены дубликаты!!!\nДля корректного объединения таблиц ,дубликаты перенесены в отдельный лист итоговой таблицы')

    # В результат объединения попадают совпадающие по ключу записи обеих таблиц и все строки из этих двух таблиц, для которых пар не нашлось. Порядок таблиц в запросе не важен.

# Создаем документ
wb = openpyxl.Workbook()
# создаем листы
ren_sheet = wb['Sheet']
ren_sheet.title = 'Таблица 1'
wb.create_sheet(title='Таблица 2', index=1)
wb.create_sheet(title='Совпадающие данные', index=2)
# Создаем листы для дубликатов
wb.create_sheet(title='Дубликаты первая таблица', index=3)
wb.create_sheet(title='Дубликаты вторая таблица', index=4)

# Проводим слияние
itog_df = pd.merge(first_df, second_df, how='outer', left_on=['ID'], right_on=['ID'],
                   indicator=True)

itog_df.to_excel('Тест.xlsx',index=False)
# Записываем каждый датафрейм в соответсвующий лист
left_df = itog_df[itog_df['_merge'] == 'left_only']
left_df.drop(['_merge'], axis=1, inplace=True)
for r in dataframe_to_rows(left_df, index=False, header=True):
    wb['Таблица 1'].append(r)

right_df = itog_df[itog_df['_merge'] == 'right_only']
right_df.drop(['_merge'], axis=1, inplace=True)
for r in dataframe_to_rows(right_df, index=False, header=True):
    wb['Таблица 2'].append(r)

both_df = itog_df[itog_df['_merge'] == 'both']
both_df.drop(['_merge'], axis=1, inplace=True)
for r in dataframe_to_rows(both_df, index=False, header=True):
    wb['Совпадающие данные'].append(r)

# Записываем дубликаты в соответствующие листы
for r in dataframe_to_rows(duplicates_first_df, index=False, header=True):
    wb['Дубликаты первая таблица'].append(r)

for r in dataframe_to_rows(duplicates_second_df, index=False, header=True):
    wb['Дубликаты вторая таблица'].append(r)

# Сохраняем
t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)
# Сохраняем итоговый файл

wb.save(f'{path_to_end_folder_comparison}/Результат слияния 2 таблиц от {current_time}.xlsx')